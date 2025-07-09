import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
from openpyxl import load_workbook


def load_login_data_xlsx():
    filepath = os.path.join(os.path.dirname(__file__), "test_data", "test_login_data.xlsx")
    wb = load_workbook(filename=filepath)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, expected, error_message, description = row
        data.append({
            "username": username if username else "",
            "password": password if password else "",
            "expected": expected.strip().lower(),
            "error_message": error_message or "",
            "description": description or ""
        })
    return data


@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    driver.maximize_window()
    time.sleep(2)
    yield driver
    driver.quit()


@pytest.mark.parametrize("data", load_login_data_xlsx())
def test_login(driver, data):
    username = data["username"]
    password = data["password"]
    expected = data["expected"]
    expected_error = data["error_message"]
    desc = data["description"]

    print(f"Test Case: {desc}")

    # Input fields
    if username:
        driver.find_element(By.NAME, "username").send_keys(username)
    if password:
        driver.find_element(By.NAME, "password").send_keys(password)

    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    time.sleep(2)

    if expected == "success":
        heading = driver.find_element(By.TAG_NAME, "h6").text
        assert "dashboard" in heading.lower(), f"Expected successful login for: {desc}"
        print("Success login test passed.")
    else:
        try:
            # Error message after failed login
            error = driver.find_element(By.XPATH, "//p[contains(@class,'oxd-alert-content-text')]").text
            assert expected_error.lower() in error.lower(), f"Error message mismatch for: {desc}"
            print("Negative login test passed.")
        except:
            # Error message for empty required fields
            username_error = driver.find_elements(By.XPATH, "//input[@name='username']/ancestor::div[contains(@class,'oxd-input-group')]//span")
            password_error = driver.find_elements(By.XPATH, "//input[@name='password']/ancestor::div[contains(@class,'oxd-input-group')]//span")

            if expected_error.lower() == "required":
                assert username_error or password_error, f"Required field error missing: {desc}"
                print("Empty field validation passed.")
            else:
                raise AssertionError(f"No matching error found for: {desc}")
